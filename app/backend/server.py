""""CRCE Lab Manager — FastAPI backend with SQLite multi-database architecture."""
from dotenv import load_dotenv
load_dotenv()
import os
import uuid
from datetime import datetime, timedelta, timezone
from typing import Optional, List

from dotenv import load_dotenv

load_dotenv()

from fastapi import FastAPI, Depends, HTTPException, APIRouter
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, EmailStr, Field

from db import init_main_db, main_db, lab_db, init_lab_db, lab_db_path, migrate_equipment_columns, migrate_lab_columns
from auth import (
    hash_password,
    verify_password,
    create_token,
    get_current_user,
    require_role,
)
from chatbot import chat_reply
from seed import run_seed
from fastapi import FastAPI, Depends, HTTPException, APIRouter, UploadFile, File
import tempfile, re
from fastapi.responses import StreamingResponse
import openpyxl, io

app = FastAPI(title="CRCE Lab Manager API")
api = APIRouter(prefix="/api")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class SignupReq(BaseModel):
    name: str
    email: EmailStr
    password: str = Field(min_length=6)
    roll_no: str
    department: str
    year: str


class LoginReq(BaseModel):
    email: EmailStr
    password: str


class CreateAssistantReq(BaseModel):
    name: str
    email: EmailStr
    password: str = Field(min_length=6)
    department: Optional[str] = None


class LabCreate(BaseModel):
    name: str
    location: str
    capacity: int = 0
    budget: float = 0
    department: Optional[str] = None
    assistant_id: Optional[int] = None
    incharge_name: Optional[str] = None  # ADD THIS
    incharge_id: Optional[int] = None

class LabBudgetUpdate(BaseModel):
    budget: float


class LabAssignAssistant(BaseModel):
    assistant_id: int


class EquipmentCreate(BaseModel):
    name: str
    category: Optional[str] = "general"
    description: Optional[str] = ""
    total_qty: int = 1
    cost: float = 0
    purchase_date: Optional[str] = None
    supplier_name: Optional[str] = None
    serial_no: Optional[str] = None
    remarks: Optional[str] = None


class EquipmentUpdate(BaseModel):
    name: Optional[str] = None
    category: Optional[str] = None
    description: Optional[str] = None
    total_qty: Optional[int] = None
    status: Optional[str] = None
    cost: Optional[float] = None
    purchase_date: Optional[str] = None
    supplier_name: Optional[str] = None
    serial_no: Optional[str] = None
    remarks: Optional[str] = None


class RequestCreate(BaseModel):
    equipment_id: int
    quantity: int = 1
    purpose: Optional[str] = ""


class MaintenanceCreate(BaseModel):
    equipment_id: int
    description: str
    cost: float = 0


class ChatReq(BaseModel):
    message: str
    session_id: Optional[str] = None

class AssistantCreate(BaseModel):
    name: str
    email: str
    password: str
    department: Optional[str] = None
    role: Optional[str] = "ASSISTANT"


@app.on_event("startup")
def _startup():
    init_main_db()
    migrate_equipment_columns()
    migrate_lab_columns()  # ADD THIS
    run_seed()


@api.get("/")
def root():
    return {"message": "CRCE Lab Manager API", "status": "ok"}


@api.post("/auth/signup")
def signup(req: SignupReq):
    with main_db() as conn:
        if conn.execute("SELECT 1 FROM users WHERE email = ?", (req.email,)).fetchone():
            raise HTTPException(status_code=400, detail="Email already registered")
        cur = conn.execute(
            "INSERT INTO users (email, password_hash, name, role, roll_no, department, year) "
            "VALUES (?, ?, ?, 'STUDENT', ?, ?, ?)",
            (req.email, hash_password(req.password), req.name, req.roll_no, req.department, req.year),
        )
        uid = cur.lastrowid
    token = create_token(uid, "STUDENT")
    return {
        "token": token,
        "user": {
            "id": uid, "name": req.name, "email": req.email, "role": "STUDENT",
            "roll_no": req.roll_no, "department": req.department, "year": req.year,
        },
    }


@api.post("/auth/login")
def login(req: LoginReq):
    with main_db() as conn:
        row = conn.execute(
            "SELECT id, email, password_hash, name, role, roll_no, department, year FROM users WHERE email = ?",
            (req.email,),
        ).fetchone()
    if not row or not verify_password(req.password, row["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    token = create_token(row["id"], row["role"])
    user = dict(row)
    user.pop("password_hash", None)
    return {"token": token, "user": user}


@api.get("/auth/me")
def me(user: dict = Depends(get_current_user)):
    return user


@api.post("/admin/create-assistant")
def create_assistant(body: AssistantCreate, _admin: dict = Depends(require_role("ADMIN"))):
    role = getattr(body, "role", "ASSISTANT") or "ASSISTANT"
    if role not in ("ASSISTANT", "INCHARGE"):
        role = "ASSISTANT"
    with main_db() as conn:
        conn.execute(
            "INSERT INTO users (email, password_hash, name, role, department) VALUES (?, ?, ?, ?, ?)",
            (body.email, hash_password(body.password), body.name, role, body.department),
        )
    return {"ok": True}

@api.get("/admin/assistants")
def list_assistants(_admin: dict = Depends(require_role("ADMIN"))):
    with main_db() as conn:
        rows = conn.execute(
            "SELECT id, email, name, role, department FROM users WHERE role IN ('ASSISTANT','INCHARGE') ORDER BY id"
        ).fetchall()
    return [dict(r) for r in rows]


@api.get("/admin/users")
def list_users(_admin: dict = Depends(require_role("ADMIN"))):
    with main_db() as conn:
        rows = conn.execute(
            "SELECT id, email, name, role, roll_no, department, year, created_at FROM users ORDER BY id"
        ).fetchall()
    return [dict(r) for r in rows]


@api.get("/labs")
def list_labs(user: dict = Depends(get_current_user)):
    with main_db() as conn:
        rows = conn.execute(
            "SELECT l.id, l.name, l.location, l.capacity, l.budget, l.department, "
            "l.db_name, l.assistant_id, l.incharge_id, l.created_at, "
            "u.name AS assistant_name, i.name AS incharge_name "
            "FROM labs l "
            "LEFT JOIN users u ON u.id = l.assistant_id "
            "LEFT JOIN users i ON i.id = l.incharge_id "
            "ORDER BY l.id"
        ).fetchall()
    return [dict(r) for r in rows]


@api.get("/labs/{lab_id}")
def get_lab(lab_id: int, user: dict = Depends(get_current_user)):
    with main_db() as conn:
        row = conn.execute(
            "SELECT l.id, l.name, l.location, l.capacity, l.budget, l.department, "
            "l.db_name, l.assistant_id, l.incharge_id, l.created_at, "
            "u.name AS assistant_name, i.name AS incharge_name "
            "FROM labs l "
            "LEFT JOIN users u ON u.id = l.assistant_id "
            "LEFT JOIN users i ON i.id = l.incharge_id "
            "ORDER BY l.id"
        ).fetchall()
    if not row:
        raise HTTPException(status_code=404, detail="Lab not found")
    return dict(row)


@api.post("/admin/labs")
def create_lab(req: LabCreate, _admin: dict = Depends(require_role("ADMIN"))):
    with main_db() as conn:
        cur = conn.execute(
            "INSERT INTO labs (name, location, capacity, budget, department, db_name, assistant_id, incharge_id) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (req.name, req.location, req.capacity, req.budget, req.department, "pending", req.assistant_id, req.incharge_id),
        )
        lab_id = cur.lastrowid
        conn.execute("UPDATE labs SET db_name = ? WHERE id = ?", (f"lab_{lab_id}.sqlite", lab_id))
    init_lab_db(lab_id)
    return {"id": lab_id, "name": req.name, "db_name": f"lab_{lab_id}.sqlite"}


@api.put("/admin/labs/{lab_id}/budget")
def set_lab_budget(lab_id: int, req: LabBudgetUpdate, _admin: dict = Depends(require_role("ADMIN"))):
    with main_db() as conn:
        if not conn.execute("SELECT 1 FROM labs WHERE id = ?", (lab_id,)).fetchone():
            raise HTTPException(status_code=404, detail="Lab not found")
        conn.execute("UPDATE labs SET budget = ? WHERE id = ?", (req.budget, lab_id))
    return {"id": lab_id, "budget": req.budget}


@api.put("/admin/labs/{lab_id}/assign-assistant")
def assign_assistant(lab_id: int, req: LabAssignAssistant, _admin: dict = Depends(require_role("ADMIN"))):
    with main_db() as conn:
        u = conn.execute("SELECT id FROM users WHERE id = ? AND role = 'ASSISTANT'", (req.assistant_id,)).fetchone()
        if not u:
            raise HTTPException(status_code=400, detail="Assistant not found")
        conn.execute("UPDATE labs SET assistant_id = ? WHERE id = ?", (req.assistant_id, lab_id))
    return {"id": lab_id, "assistant_id": req.assistant_id}

class LabAssignIncharge(BaseModel):
    incharge_id: int

@api.put("/admin/labs/{lab_id}/assign-incharge")
def assign_incharge(lab_id: int, req: LabAssignIncharge, _admin: dict = Depends(require_role("ADMIN"))):
    with main_db() as conn:
        u = conn.execute(
            "SELECT id FROM users WHERE id = ? AND role = 'INCHARGE'",
            (req.incharge_id,)
        ).fetchone()
        if not u:
            raise HTTPException(status_code=400, detail="Incharge user not found")
        conn.execute(
            "UPDATE labs SET incharge_id = ? WHERE id = ?",
            (req.incharge_id, lab_id)
        )
    return {"id": lab_id, "incharge_id": req.incharge_id}

@api.delete("/admin/labs/{lab_id}")
def delete_lab(lab_id: int, _admin: dict = Depends(require_role("ADMIN"))):
    with main_db() as conn:
        conn.execute("DELETE FROM labs WHERE id = ?", (lab_id,))
    p = lab_db_path(lab_id)
    if p.exists():
        try:
            p.unlink()
        except Exception:
            pass
    return {"deleted": lab_id}

@api.post("/admin/labs/{lab_id}/import-registry")
async def import_registry(lab_id: int, file: UploadFile = File(...), _admin: dict = Depends(require_role("ADMIN"))):
    import pdfplumber, openpyxl
    suffix = file.filename.split(".")[-1].lower()
    contents = await file.read()
    rows = []

    with tempfile.NamedTemporaryFile(suffix=f".{suffix}", delete=False) as tmp:
        tmp.write(contents)
        tmp_path = tmp.name

    if suffix == "pdf":
        with pdfplumber.open(tmp_path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    for row in table[1:]:  # skip header
                        if not row or not row[1]:
                            continue
                        try:
                            rows.append({
                                "name": str(row[1]).strip(),
                                "total_qty": int(re.sub(r"[^\d]", "", str(row[2])) or 1),
                                "purchase_date": str(row[3]).strip() if row[3] else None,
                                "supplier_name": str(row[4]).strip() if row[4] else None,
                                "serial_no": str(row[5]).strip() if row[5] else None,
                                "cost": float(re.sub(r"[^\d.]", "", str(row[6])) or 0),
                                "remarks": str(row[7]).strip() if len(row) > 7 and row[7] else None,
                            })
                        except:
                            continue

    elif suffix in ("xlsx", "xls"):
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        ws = wb.active
        for row in list(ws.iter_rows(values_only=True))[1:]:
            if not row[1]:
                continue
            try:
                rows.append({
                    "name": str(row[1]).strip(),
                    "total_qty": int(row[2] or 1),
                    "purchase_date": str(row[3]) if row[3] else None,
                    "supplier_name": str(row[4]) if row[4] else None,
                    "serial_no": str(row[5]) if row[5] else None,
                    "cost": float(row[6] or 0),
                    "remarks": str(row[7]) if len(row) > 7 and row[7] else None,
                })
            except:
                continue
    else:
        raise HTTPException(status_code=400, detail="Only PDF or Excel files supported")

    with lab_db(lab_id) as conn:
        for r in rows:
            conn.execute(
                "INSERT INTO equipment (name, total_qty, available_qty, cost, purchase_date, supplier_name, serial_no, remarks, status) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'AVAILABLE')",
                (r["name"], r["total_qty"], r["total_qty"], r["cost"],
                 r["purchase_date"], r["supplier_name"], r["serial_no"], r["remarks"])
            )
    return {"imported": len(rows)}

@api.get("/admin/labs/{lab_id}/export")
def export_lab_registry(lab_id: int, _admin: dict = Depends(require_role("ADMIN"))):
    with main_db() as conn:
        lab = conn.execute("SELECT * FROM labs WHERE id = ?", (lab_id,)).fetchone()
    if not lab:
        raise HTTPException(status_code=404, detail="Lab not found")
    with lab_db(lab_id) as conn:
        items = conn.execute("SELECT * FROM equipment ORDER BY id").fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = lab["name"]
    ws.append(["Sr.No", "Description", "Quantity", "Purchase Date", 
                "Supplier", "Serial No", "Cost (₹)", "Status", "Remarks"])
    for i, it in enumerate(items, 1):
        ws.append([i, it["name"], it["total_qty"], it["purchase_date"],
                   it["supplier_name"], it["serial_no"], it["cost"],
                   it["status"], it["remarks"]])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{lab['name'].replace(' ', '_')}_registry.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})

def _ensure_assistant_for_lab(lab_id: int, user: dict):
    with main_db() as conn:
        lab = conn.execute("SELECT * FROM labs WHERE id = ?", (lab_id,)).fetchone()
    if not lab:
        raise HTTPException(status_code=404, detail="Lab not found")
    if user["role"] == "ADMIN":
        return
    if user["role"] == "INCHARGE" and lab["incharge_id"] == user["id"]:
        return
    if user["role"] == "ASSISTANT" and lab["assistant_id"] == user["id"]:
        return
    raise HTTPException(status_code=403, detail="Not your lab")

@api.get("/labs/{lab_id}/equipment")
def list_equipment(lab_id: int, user: dict = Depends(get_current_user)):
    with main_db() as conn:
        if not conn.execute("SELECT 1 FROM labs WHERE id = ?", (lab_id,)).fetchone():
            raise HTTPException(status_code=404, detail="Lab not found")
    with lab_db(lab_id) as conn:
        rows = conn.execute("SELECT * FROM equipment ORDER BY id DESC").fetchall()
    return [dict(r) for r in rows]


@api.post("/labs/{lab_id}/equipment")
def add_equipment(lab_id: int, req: EquipmentCreate, user: dict = Depends(get_current_user)):
    _ensure_assistant_for_lab(lab_id, user)
    with lab_db(lab_id) as conn:
        cur = conn.execute(
            "INSERT INTO equipment (name, category, description, total_qty, available_qty, cost, status) "
            "VALUES (?, ?, ?, ?, ?, ?, 'AVAILABLE')",
            (req.name, req.category, req.description, req.total_qty, req.total_qty, req.cost),
        )
        eid = cur.lastrowid
        row = conn.execute("SELECT * FROM equipment WHERE id = ?", (eid,)).fetchone()
    return dict(row)


@api.put("/labs/{lab_id}/equipment/{eid}")
def update_equipment(lab_id: int, eid: int, req: EquipmentUpdate, user: dict = Depends(get_current_user)):
    _ensure_assistant_for_lab(lab_id, user)
    fields = {k: v for k, v in req.model_dump().items() if v is not None}
    if not fields:
        raise HTTPException(status_code=400, detail="No fields to update")
    with lab_db(lab_id) as conn:
        existing = conn.execute("SELECT * FROM equipment WHERE id = ?", (eid,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Equipment not found")
        if "total_qty" in fields:
            issued = existing["total_qty"] - existing["available_qty"]
            new_total = max(fields["total_qty"], issued)
            fields["total_qty"] = new_total
            fields["available_qty"] = max(0, new_total - issued)
        sets = ", ".join(f"{k} = ?" for k in fields.keys())
        vals = list(fields.values()) + [eid]
        conn.execute(f"UPDATE equipment SET {sets} WHERE id = ?", vals)
        row = conn.execute("SELECT * FROM equipment WHERE id = ?", (eid,)).fetchone()
    return dict(row)


@api.delete("/labs/{lab_id}/equipment/{eid}")
def delete_equipment(lab_id: int, eid: int, user: dict = Depends(get_current_user)):
    _ensure_assistant_for_lab(lab_id, user)
    with lab_db(lab_id) as conn:
        conn.execute("DELETE FROM equipment WHERE id = ?", (eid,))
    return {"deleted": eid}


@api.post("/labs/{lab_id}/requests")
def create_request(lab_id: int, req: RequestCreate, user: dict = Depends(require_role("STUDENT"))):
    with main_db() as conn:
        if not conn.execute("SELECT 1 FROM labs WHERE id = ?", (lab_id,)).fetchone():
            raise HTTPException(status_code=404, detail="Lab not found")
    with lab_db(lab_id) as conn:
        eq = conn.execute("SELECT * FROM equipment WHERE id = ?", (req.equipment_id,)).fetchone()
        if not eq:
            raise HTTPException(status_code=404, detail="Equipment not found")
        if eq["available_qty"] < req.quantity:
            raise HTTPException(status_code=400, detail="Not enough quantity available")
        cur = conn.execute(
            "INSERT INTO requests (equipment_id, student_id, student_name, student_email, quantity, purpose, status) "
            "VALUES (?, ?, ?, ?, ?, ?, 'PENDING')",
            (req.equipment_id, user["id"], user["name"], user["email"], req.quantity, req.purpose),
        )
        rid = cur.lastrowid
        row = conn.execute("SELECT * FROM requests WHERE id = ?", (rid,)).fetchone()
    return dict(row)


@api.get("/labs/{lab_id}/requests")
def list_requests(lab_id: int, user: dict = Depends(get_current_user)):
    with lab_db(lab_id) as conn:
        if user["role"] == "STUDENT":
            rows = conn.execute(
                "SELECT r.*, e.name AS equipment_name FROM requests r "
                "LEFT JOIN equipment e ON e.id = r.equipment_id WHERE r.student_id = ? ORDER BY r.id DESC",
                (user["id"],),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT r.*, e.name AS equipment_name FROM requests r "
                "LEFT JOIN equipment e ON e.id = r.equipment_id ORDER BY r.id DESC"
            ).fetchall()
    return [dict(r) for r in rows]


@api.put("/labs/{lab_id}/requests/{rid}/approve")
def approve_request(lab_id: int, rid: int, user: dict = Depends(get_current_user)):
    _ensure_assistant_for_lab(lab_id, user)
    with lab_db(lab_id) as conn:
        r = conn.execute("SELECT * FROM requests WHERE id = ?", (rid,)).fetchone()
        if not r:
            raise HTTPException(status_code=404, detail="Request not found")
        if r["status"] != "PENDING":
            raise HTTPException(status_code=400, detail="Request not pending")
        eq = conn.execute("SELECT * FROM equipment WHERE id = ?", (r["equipment_id"],)).fetchone()
        if eq["available_qty"] < r["quantity"]:
            raise HTTPException(status_code=400, detail="Not enough available")
        due = (datetime.now(timezone.utc) + timedelta(days=7)).isoformat()
        conn.execute(
            "UPDATE requests SET status = 'ISSUED', approved_at = datetime('now'), due_date = ? WHERE id = ?",
            (due, rid),
        )
        conn.execute(
            "UPDATE equipment SET available_qty = available_qty - ? WHERE id = ?",
            (r["quantity"], r["equipment_id"]),
        )
    return {"id": rid, "status": "ISSUED", "due_date": due}


@api.put("/labs/{lab_id}/requests/{rid}/reject")
def reject_request(lab_id: int, rid: int, user: dict = Depends(get_current_user)):
    _ensure_assistant_for_lab(lab_id, user)
    with lab_db(lab_id) as conn:
        r = conn.execute("SELECT * FROM requests WHERE id = ?", (rid,)).fetchone()
        if not r or r["status"] != "PENDING":
            raise HTTPException(status_code=400, detail="Cannot reject")
        conn.execute("UPDATE requests SET status = 'REJECTED' WHERE id = ?", (rid,))
    return {"id": rid, "status": "REJECTED"}


@api.put("/labs/{lab_id}/requests/{rid}/return")
def return_request(lab_id: int, rid: int, user: dict = Depends(get_current_user)):
    _ensure_assistant_for_lab(lab_id, user)
    with lab_db(lab_id) as conn:
        r = conn.execute("SELECT * FROM requests WHERE id = ?", (rid,)).fetchone()
        if not r or r["status"] != "ISSUED":
            raise HTTPException(status_code=400, detail="Not an issued request")
        conn.execute(
            "UPDATE requests SET status = 'RETURNED', returned_at = datetime('now') WHERE id = ?",
            (rid,),
        )
        conn.execute(
            "UPDATE equipment SET available_qty = available_qty + ? WHERE id = ?",
            (r["quantity"], r["equipment_id"]),
        )
    return {"id": rid, "status": "RETURNED"}


@api.get("/labs/{lab_id}/maintenance")
def list_maintenance(lab_id: int, user: dict = Depends(get_current_user)):
    with lab_db(lab_id) as conn:
        rows = conn.execute(
            "SELECT m.*, e.name AS equipment_name FROM maintenance m "
            "LEFT JOIN equipment e ON e.id = m.equipment_id ORDER BY m.id DESC"
        ).fetchall()
    return [dict(r) for r in rows]


@api.post("/labs/{lab_id}/maintenance")
def add_maintenance(lab_id: int, req: MaintenanceCreate, user: dict = Depends(get_current_user)):
    _ensure_assistant_for_lab(lab_id, user)
    with lab_db(lab_id) as conn:
        eq = conn.execute("SELECT * FROM equipment WHERE id = ?", (req.equipment_id,)).fetchone()
        if not eq:
            raise HTTPException(status_code=404, detail="Equipment not found")
        cur = conn.execute(
            "INSERT INTO maintenance (equipment_id, description, cost, status) VALUES (?, ?, ?, 'IN_PROGRESS')",
            (req.equipment_id, req.description, req.cost),
        )
        mid = cur.lastrowid
        conn.execute("UPDATE equipment SET status = 'MAINTENANCE' WHERE id = ?", (req.equipment_id,))
    return {"id": mid, "status": "IN_PROGRESS"}


@api.put("/labs/{lab_id}/maintenance/{mid}/complete")
def complete_maintenance(lab_id: int, mid: int, user: dict = Depends(get_current_user)):
    _ensure_assistant_for_lab(lab_id, user)
    with lab_db(lab_id) as conn:
        m = conn.execute("SELECT * FROM maintenance WHERE id = ?", (mid,)).fetchone()
        if not m:
            raise HTTPException(status_code=404, detail="Not found")
        conn.execute(
            "UPDATE maintenance SET status = 'COMPLETED', completed_at = datetime('now') WHERE id = ?",
            (mid,),
        )
        conn.execute("UPDATE equipment SET status = 'AVAILABLE' WHERE id = ?", (m["equipment_id"],))
    return {"id": mid, "status": "COMPLETED"}


@api.get("/students/me/borrowed")
def my_borrowed(user: dict = Depends(require_role("STUDENT"))):
    out = []
    with main_db() as conn:
        labs = conn.execute("SELECT id, name FROM labs").fetchall()
    for lab in labs:
        try:
            with lab_db(lab["id"]) as conn:
                rows = conn.execute(
                    "SELECT r.*, e.name AS equipment_name FROM requests r "
                    "LEFT JOIN equipment e ON e.id = r.equipment_id "
                    "WHERE r.student_id = ? ORDER BY r.id DESC",
                    (user["id"],),
                ).fetchall()
            for r in rows:
                d = dict(r)
                d["lab_id"] = lab["id"]
                d["lab_name"] = lab["name"]
                out.append(d)
        except Exception:
            continue
    return out


@api.get("/admin/analytics")
def analytics(_admin: dict = Depends(require_role("ADMIN"))):
    with main_db() as conn:
        labs = conn.execute("SELECT id, name, budget FROM labs").fetchall()
        users = conn.execute("SELECT role, COUNT(*) AS c FROM users GROUP BY role").fetchall()
    per_lab = []
    total_equipment = 0
    total_issued = 0
    total_pending_requests = 0
    for lab in labs:
        try:
            with lab_db(lab["id"]) as conn:
                eq = conn.execute(
                    "SELECT COUNT(*) AS c, COALESCE(SUM(total_qty),0) AS total, "
                    "COALESCE(SUM(total_qty - available_qty),0) AS issued, "
                    "COALESCE(SUM(cost*total_qty),0) AS spend FROM equipment"
                ).fetchone()
                pending = conn.execute(
                    "SELECT COUNT(*) AS c FROM requests WHERE status = 'PENDING'"
                ).fetchone()["c"]
            per_lab.append({
                "lab_id": lab["id"], "lab_name": lab["name"], "budget": lab["budget"],
                "spend": eq["spend"], "equipment_count": eq["c"],
                "total_units": eq["total"], "issued_units": eq["issued"],
                "pending_requests": pending,
            })
            total_equipment += eq["total"]
            total_issued += eq["issued"]
            total_pending_requests += pending
        except Exception:
            continue
    return {
        "users": {row["role"]: row["c"] for row in users},
        "totals": {
            "labs": len(labs),
            "equipment_units": total_equipment,
            "issued_units": total_issued,
            "pending_requests": total_pending_requests,
        },
        "per_lab": per_lab,
    }


@api.post("/chatbot")
async def chatbot(req: ChatReq, user: dict = Depends(get_current_user)):
    sid = req.session_id or f"user-{user['id']}-{uuid.uuid4().hex[:8]}"
    try:
        reply = await chat_reply(sid, req.message, user_name=user["name"])
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Chatbot error: {e}")
    return {"session_id": sid, "reply": reply}


app.include_router(api)
